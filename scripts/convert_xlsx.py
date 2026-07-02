#!/usr/bin/env python3
"""Convert the festival "Бессонница 2026" Excel timetables into a single
normalized program.json used by the offline web app.

Two source layouts are supported and merged:
  * Animation (night screenings): Время | Экран    | Программа | Фильмы   | Возраст
  * Non-animation (day program):  Время | Площадка | Событие   | Описание | Возраст

Events routinely cross midnight (e.g. an entry listed under "9 июля" that
starts at 00:30 actually happens in the early morning of 10 July). Real
start/end datetimes are computed so the app can sort and show "now"
correctly. The rule: within a day sheet the festival "night" runs from the
evening into the small hours, and nothing is scheduled between ~04:00 and
~09:00, so any time with hour < 9 belongs to the next calendar day.
"""
import json
import re
import sys
import hashlib
from pathlib import Path
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SOURCES = [
    (ROOT / "source_xlsx" / "animation.xlsx", "animation"),
    (ROOT / "source_xlsx" / "nonanimation.xlsx", "program"),
]
OUT = ROOT / "data" / "program.json"

MONTHS = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5, "июня": 6,
    "июля": 7, "августа": 8, "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}
YEAR = 2026
# Hour threshold: times earlier than this belong to the following calendar day.
NIGHT_ROLLOVER_HOUR = 9


def parse_sheet_date(sheet_name, title_cell):
    """Return (date, weekday_iso) from a sheet name like '9 июля'."""
    text = f"{sheet_name} {title_cell or ''}"
    m = re.search(r"(\d{1,2})\s+([а-яё]+)", text.lower())
    if not m:
        return None
    day = int(m.group(1))
    month = MONTHS.get(m.group(2))
    if not month:
        return None
    return datetime(YEAR, month, day)


def parse_time_range(raw):
    """Return (start 'HH:MM', end 'HH:MM'|None) from '21:20–22:30'."""
    if not raw:
        return None, None
    s = str(raw).strip().replace("–", "-").replace("—", "-").replace("—", "-")
    parts = [p.strip() for p in s.split("-") if p.strip()]
    times = []
    for p in parts:
        m = re.match(r"^(\d{1,2}):(\d{2})$", p)
        if m:
            times.append(f"{int(m.group(1)):02d}:{m.group(2)}")
    if not times:
        return None, None
    start = times[0]
    end = times[1] if len(times) > 1 else None
    return start, end


def to_datetime(base_date, hhmm):
    """Attach hh:mm to base_date, rolling into the next day for small hours."""
    if not hhmm:
        return None
    h, m = int(hhmm[:2]), int(hhmm[3:])
    dt = base_date.replace(hour=h, minute=m)
    if h < NIGHT_ROLLOVER_HOUR:
        dt += timedelta(days=1)
    return dt


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def normalize_venue(v):
    """The source table leaves some stage names as keyboard-mash placeholders
    (e.g. 'тстцтсттсцтс'). Show a readable 'stage TBD' label instead."""
    s = clean(v)
    letters = re.sub(r"[^а-яёa-z]", "", s.lower())
    if letters and set(letters) <= set("тсц"):
        return "Сцена (уточняется)"
    return s


def split_films(text):
    if not text:
        return []
    return [p.strip() for p in str(text).split(",") if p.strip()]


def make_id(*parts):
    """FNV-1a 32-bit over UTF-8. Mirrored byte-for-byte in app.js so that
    favorites keep matching when the user re-imports an updated table."""
    data = "|".join(parts).encode("utf-8")
    h = 2166136261
    for b in data:
        h ^= b
        h = (h * 16777619) & 0xFFFFFFFF
    return format(h, "08x")


def convert():
    events = []
    days = {}
    for path, kind in SOURCES:
        if not path.exists():
            print(f"warning: {path} missing, skipping", file=sys.stderr)
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            title_cell = clean(rows[0][0]) if rows[0] else ""
            base_date = parse_sheet_date(sheet_name, title_cell)
            if base_date is None:
                continue
            date_iso = base_date.strftime("%Y-%m-%d")
            days.setdefault(date_iso, {"date": date_iso, "label": sheet_name.strip()})
            # rows[0] = section title, rows[1] = header, rows[2:] = data
            for r in rows[2:]:
                if not r:
                    continue
                time_raw = clean(r[0]) if len(r) > 0 else ""
                col_place = normalize_venue(r[1]) if len(r) > 1 else ""
                col_title = clean(r[2]) if len(r) > 2 else ""
                col_desc = clean(r[3]) if len(r) > 3 else ""
                col_age = clean(r[4]) if len(r) > 4 else ""
                start, end = parse_time_range(time_raw)
                if not start or not col_title:
                    continue
                start_dt = to_datetime(base_date, start)
                end_dt = to_datetime(base_date, end) if end else None
                if end_dt and start_dt and end_dt <= start_dt:
                    end_dt += timedelta(days=1)
                films = split_films(col_desc) if kind == "animation" else []
                description = "" if kind == "animation" else col_desc
                ev = {
                    "id": make_id(kind, date_iso, start, col_place, col_title),
                    "type": kind,
                    "date": date_iso,
                    "start": start,
                    "end": end,
                    "startISO": start_dt.strftime("%Y-%m-%dT%H:%M:00") if start_dt else None,
                    "endISO": end_dt.strftime("%Y-%m-%dT%H:%M:00") if end_dt else None,
                    "venue": col_place,
                    "title": col_title,
                    "description": description,
                    "films": films,
                    "age": col_age,
                }
                events.append(ev)

    events.sort(key=lambda e: (e["startISO"] or "", e["venue"]))
    day_list = sorted(days.values(), key=lambda d: d["date"])
    payload = {
        "festival": "Бессонница 2026",
        "year": YEAR,
        "generatedAt": None,  # filled by caller / build; keep deterministic here
        "version": 1,
        "days": day_list,
        "venues": sorted({e["venue"] for e in events if e["venue"]}),
        "events": events,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"Wrote {len(events)} events across {len(day_list)} days -> {OUT}")


if __name__ == "__main__":
    convert()
